"""Deterministic five-sheet workbook writer; no network or inference."""
from __future__ import annotations

import os
import re
import tempfile
from io import BytesIO
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from pipeline.cloud.claude_client import AnalysisResult
from pipeline.extraction.schemas import ExtractionPayload

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
FMT_NUMBER = "#,##0.00"
FMT_PERCENT = "0.0%"
PERCENT_COLUMNS: dict[str, set[str]] = {"Variance": {"delta_pct"}}
SUMMARY_PERCENT_RE = re.compile(r"(margin|ratio|growth|rate|pct|%)", re.IGNORECASE)
CORE_TIMESTAMP_RE = re.compile(
    rb"(<dcterms:(?:created|modified)\b[^>]*>)[^<]*"
    rb"(</dcterms:(?:created|modified)>)"
)
FIXED_OOXML_TIMESTAMP = b"2026-01-01T00:00:00Z"

SUMMARY_COLUMNS = ["name", "value", "period", "formula_used", "inputs"]
VARIANCE_COLUMNS = [
    "metric", "period_a", "period_b", "delta_abs", "delta_pct", "commentary"
]
ADJUSTMENT_COLUMNS = [
    "description", "metric_affected", "period", "original_value",
    "adjusted_value", "rationale", "inputs",
]
RAW_FIGURE_COLUMNS = [
    "figure_id", "label", "value", "unit", "period", "statement",
    "source_page", "verbatim_context",
]


def output_filename(stem: str, sha12: str) -> str:
    return f"{stem}.{sha12}.xlsx"


def _frames(
    result: AnalysisResult, payload: ExtractionPayload
) -> dict[str, pd.DataFrame]:
    summary_rows = []
    for metric in result.computed_metrics:
        row = metric.model_dump()
        row["inputs"] = ", ".join(metric.inputs)
        summary_rows.append(row)
    adjustment_rows = []
    for adjustment in result.adjustments:
        row = adjustment.model_dump()
        row["inputs"] = ", ".join(adjustment.inputs)
        adjustment_rows.append(row)
    raw_rows = []
    for figure in payload.figures:
        data = figure.model_dump(mode="json")
        raw_rows.append({column: data.get(column) for column in RAW_FIGURE_COLUMNS})
    quality_rows = (
        [{"flag": flag} for flag in result.data_quality_flags]
        if result.data_quality_flags
        else [{"flag": "No issues flagged"}]
    )
    return {
        "Summary": pd.DataFrame(summary_rows, columns=SUMMARY_COLUMNS),
        "Variance": pd.DataFrame(
            [item.model_dump() for item in result.variance_analysis],
            columns=VARIANCE_COLUMNS,
        ),
        "Adjustments": pd.DataFrame(adjustment_rows, columns=ADJUSTMENT_COLUMNS),
        "Raw Figures": pd.DataFrame(raw_rows, columns=RAW_FIGURE_COLUMNS),
        "Data Quality": pd.DataFrame(quality_rows, columns=["flag"]),
    }


def write_workbook(
    result: AnalysisResult, payload: ExtractionPayload, out_path: Path
) -> Path:
    """Write and atomically publish the fixed five-sheet workbook."""
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    frames = _frames(result, payload)
    temporary: str | None = None
    try:
        with tempfile.NamedTemporaryFile(
            suffix=".xlsx", dir=out_path.parent, delete=False
        ) as handle:
            temporary = handle.name
        with pd.ExcelWriter(temporary, engine="openpyxl") as writer:
            for name, frame in frames.items():
                frame.to_excel(writer, sheet_name=name, index=False)
            for name in frames:
                _format_sheet(writer.sheets[name], PERCENT_COLUMNS.get(name, set()))
            _apply_summary_percent(writer.sheets["Summary"])
        _normalize_xlsx(Path(temporary))
        os.replace(temporary, out_path)
        return out_path
    finally:
        if temporary and Path(temporary).exists():
            Path(temporary).unlink(missing_ok=True)


def _normalize_xlsx(path: Path) -> None:
    """Remove OOXML save-time entropy before the atomic publish step."""
    with ZipFile(path, "r") as source:
        entries = {name: source.read(name) for name in source.namelist()}
    core_name = "docProps/core.xml"
    if core_name in entries:
        entries[core_name] = CORE_TIMESTAMP_RE.sub(
            rb"\g<1>" + FIXED_OOXML_TIMESTAMP + rb"\g<2>",
            entries[core_name],
        )
    output = BytesIO()
    with ZipFile(output, "w", compression=ZIP_DEFLATED, compresslevel=9) as target:
        for name in sorted(entries):
            info = ZipInfo(name, date_time=(2026, 1, 1, 0, 0, 0))
            info.compress_type = ZIP_DEFLATED
            info.create_system = 3
            info.external_attr = 0o600 << 16
            target.writestr(info, entries[name])
    path.write_bytes(output.getvalue())


def _format_sheet(ws: Worksheet, percent_cols: set[str]) -> None:
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    headers = {cell.column: str(cell.value) if cell.value is not None else "" for cell in ws[1]}
    for column_index, header in headers.items():
        data_cells = [ws.cell(row=row, column=column_index) for row in range(2, ws.max_row + 1)]
        if header in percent_cols:
            for cell in data_cells:
                cell.number_format = FMT_PERCENT
        elif any(
            isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool)
            for cell in data_cells
        ):
            for cell in data_cells:
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    cell.number_format = FMT_NUMBER
    _autowidth(ws)


def _apply_summary_percent(ws: Worksheet) -> None:
    for row in range(2, ws.max_row + 1):
        metric_name = ws.cell(row=row, column=1).value
        if metric_name is not None and SUMMARY_PERCENT_RE.search(str(metric_name)):
            ws.cell(row=row, column=2).number_format = FMT_PERCENT


def _autowidth(ws: Worksheet, min_width: int = 10, max_width: int = 60) -> None:
    for index in range(1, ws.max_column + 1):
        values = [
            len(str(ws.cell(row=row, column=index).value))
            for row in range(1, ws.max_row + 1)
            if ws.cell(row=row, column=index).value is not None
        ]
        desired = (max(values) if values else 0) + 2
        width = min(max(desired, min_width), max_width)
        ws.column_dimensions[get_column_letter(index)].width = width
        if desired >= max_width:
            for row in range(1, ws.max_row + 1):
                cell = ws.cell(row=row, column=index)
                cell.alignment = cell.alignment.copy(wrap_text=True)
