"""Generate deterministic, non-sensitive document fixtures."""
from __future__ import annotations

import re
from pathlib import Path
from datetime import datetime, timezone
from io import BytesIO
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

import fitz
from openpyxl import Workbook

ROOT = Path(__file__).resolve().parent


def _native_pdf() -> None:
    document = fitz.open()
    page = document.new_page(width=612, height=792)
    page.insert_text(
        (72, 72),
        "Acme Corporation Annual Report FY2025\nIncome Statement\n"
        "Revenue 1,250\nOperating income 250\nTotal assets 4,500\n",
        fontsize=12,
    )
    page2 = document.new_page(width=612, height=792)
    page2.insert_text(
        (72, 72),
        "Balance Sheet FY2025\nCash 500\nCurrent liabilities 400\nTotal equity 2,000\n",
        fontsize=12,
    )
    document.set_metadata(
        {"creationDate": "D:20260101000000Z", "modDate": "D:20260101000000Z"}
    )
    document.save(
        ROOT / "native_report.pdf", garbage=4, deflate=True, no_new_id=True
    )
    document.close()


def _scanned_pdf() -> None:
    source = fitz.open()
    page = source.new_page(width=612, height=792)
    page.insert_text((72, 72), "Scanned Financial Report\nRevenue 1,250\nNet income 125", fontsize=16)
    pixmap = page.get_pixmap(dpi=150, alpha=False)
    image_bytes = pixmap.tobytes("png")
    source.close()

    scanned = fitz.open()
    target = scanned.new_page(width=612, height=792)
    target.insert_image(target.rect, stream=image_bytes)
    scanned.set_metadata(
        {"creationDate": "D:20260101000000Z", "modDate": "D:20260101000000Z"}
    )
    scanned.save(
        ROOT / "scanned_report.pdf", garbage=4, deflate=True, no_new_id=True
    )
    scanned.close()


def _workbook() -> None:
    workbook = Workbook()
    income = workbook.active
    income.title = "Income"
    income.append(["Metric", "FY2024", "FY2025"])
    income.append(["Revenue", 1000, 1250])
    income.append(["Operating income", 180, 250])
    empty = workbook.create_sheet("Empty")
    empty["A1"] = None
    balance = workbook.create_sheet("Balance")
    balance.append(["Metric", "FY2025"])
    balance.append(["Total assets", 4500])
    balance.append(["Total equity", 2000])
    fixed = datetime(2026, 1, 1, tzinfo=timezone.utc)
    workbook.properties.created = fixed
    workbook.properties.modified = fixed
    destination = ROOT / "financials.xlsx"
    workbook.save(destination)
    _normalize_zip(destination)


def _normalize_zip(path: Path) -> None:
    """Rewrite OOXML with fixed entry order/timestamps for byte determinism."""
    with ZipFile(path, "r") as source:
        entries = {name: source.read(name) for name in source.namelist()}
    # openpyxl replaces the workbook's modified property with the save time,
    # even when the property was explicitly assigned before save.
    core_name = "docProps/core.xml"
    entries[core_name] = re.sub(
        rb"(<dcterms:modified\b[^>]*>)[^<]*(</dcterms:modified>)",
        rb"\g<1>2026-01-01T00:00:00Z\g<2>",
        entries[core_name],
        count=1,
    )
    output = BytesIO()
    with ZipFile(output, "w", compression=ZIP_DEFLATED, compresslevel=9) as target:
        for name in sorted(entries):
            info = ZipInfo(name, date_time=(2026, 1, 1, 0, 0, 0))
            info.compress_type = ZIP_DEFLATED
            info.create_system = 3
            info.external_attr = 0o600 << 16
            target.writestr(info, entries[name])
    path.write_bytes(output.getvalue())


def _csv() -> None:
    (ROOT / "financials.csv").write_text(
        "Metric,FY2024,FY2025\nRevenue,1000,1250\nNet income,100,125\n",
        encoding="utf-8",
    )


def _encrypted_pdf() -> None:
    # MuPDF intentionally salts newly encrypted files, which would make this
    # fixture's content hash change on every generation. This minimal PDF has
    # a fixed Standard Security Handler dictionary. Detection only needs a
    # valid encrypted container (`Document.needs_pass`); no test authenticates
    # or reads encrypted content.
    objects = [
        b"1 0 obj\n<< /Type /Catalog /Pages 2 0 R >>\nendobj\n",
        b"2 0 obj\n<< /Type /Pages /Kids [3 0 R] /Count 1 >>\nendobj\n",
        b"3 0 obj\n<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] >>\nendobj\n",
    ]
    parts = [b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n"]
    offsets: list[int] = []
    for obj in objects:
        offsets.append(sum(len(part) for part in parts))
        parts.append(obj)
    xref_offset = sum(len(part) for part in parts)
    xref_rows = [b"0000000000 65535 f \n"] + [
        f"{offset:010d} 00000 n \n".encode("ascii") for offset in offsets
    ]
    fixed_id = b"00112233445566778899AABBCCDDEEFF"
    owner_key = b"00" * 32
    user_key = b"11" * 32
    parts.extend(
        [
            b"xref\n0 4\n",
            *xref_rows,
            b"trailer\n<< /Size 4 /Root 1 0 R /ID [<"
            + fixed_id
            + b"><"
            + fixed_id
            + b">] /Encrypt << /Filter /Standard /V 1 /R 2 /P -4 /O <"
            + owner_key
            + b"> /U <"
            + user_key
            + b"> >> >>\nstartxref\n"
            + str(xref_offset).encode("ascii")
            + b"\n%%EOF\n",
        ]
    )
    (ROOT / "encrypted.pdf").write_bytes(b"".join(parts))


def main() -> None:
    ROOT.mkdir(parents=True, exist_ok=True)
    _native_pdf()
    _scanned_pdf()
    _workbook()
    _csv()
    _encrypted_pdf()


if __name__ == "__main__":
    main()
