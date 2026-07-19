from __future__ import annotations

from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook

from pipeline.cloud.claude_client import (
    AdjustmentItem,
    AnalysisResult,
    ComputedMetric,
    VarianceItem,
)
from pipeline.extraction.schemas import ExtractionPayload, FinancialFigure, StatementType
from pipeline.output.excel_writer import FMT_NUMBER, FMT_PERCENT, output_filename, write_workbook


def _result() -> AnalysisResult:
    return AnalysisResult(
        computed_metrics=[
            ComputedMetric(name="Gross Margin", value=0.42, formula_used="gross_profit / revenue",
                           inputs=["F0001", "F0002"], period="FY2025"),
            ComputedMetric(name="Working Capital", value=100.0,
                           formula_used="current_assets - current_liabilities",
                           inputs=["F0003", "F0004"], period="FY2025"),
        ],
        variance_analysis=[
            VarianceItem(metric="Revenue", period_a="FY2024", period_b="FY2025",
                         delta_abs=100.0, delta_pct=0.083, commentary="Revenue increased.")
        ],
        adjustments=[
            AdjustmentItem(description="Remove restructuring", metric_affected="Operating income",
                           period="FY2025", original_value=20, adjusted_value=30,
                           rationale="Explicit one-time item.", inputs=["F0005"])
        ],
        data_quality_flags=[],
    )


def _payload() -> ExtractionPayload:
    return ExtractionPayload(
        company="Acme", doc_type="annual report", currency_default="USD", periods=["FY2025"],
        figures=[
            FinancialFigure(figure_id="F0001", label="Revenue", value=1000, unit="thousands",
                            currency="USD", period="FY2025", statement=StatementType.income_statement,
                            source_page=1, verbatim_context="Revenue 1,000 local audit snippet")
        ], warnings=[]
    )


def test_output_filename_is_deterministic() -> None:
    assert output_filename("acme", "a" * 12) == f"acme.{ 'a' * 12 }.xlsx"


def test_workbook_sheets_values_and_formats(tmp_path: Path) -> None:
    path = write_workbook(_result(), _payload(), tmp_path / "result.xlsx")
    workbook = load_workbook(path)
    assert workbook.sheetnames == ["Summary", "Variance", "Adjustments", "Raw Figures", "Data Quality"]
    summary = workbook["Summary"]
    assert summary["B2"].value == 0.42
    assert summary["B2"].number_format == FMT_PERCENT
    assert summary["B3"].value == 100.0
    assert summary["B3"].number_format == FMT_NUMBER
    variance = workbook["Variance"]
    assert variance["E2"].value == 0.083
    assert variance["E2"].number_format == FMT_PERCENT
    assert summary.freeze_panes == "A2"
    assert summary.auto_filter.ref == summary.dimensions
    raw = workbook["Raw Figures"]
    assert raw["A2"].value == "F0001"
    assert "local audit snippet" in raw["H2"].value
    assert workbook["Data Quality"]["A2"].value == "No issues flagged"


def test_workbook_bytes_are_deterministic(tmp_path: Path) -> None:
    first = write_workbook(_result(), _payload(), tmp_path / "first.xlsx")
    second = write_workbook(_result(), _payload(), tmp_path / "second.xlsx")
    assert first.read_bytes() == second.read_bytes()
    with ZipFile(first) as archive:
        assert all(
            item.date_time == (2026, 1, 1, 0, 0, 0)
            for item in archive.infolist()
        )
        core = archive.read("docProps/core.xml")
    assert core.count(b"2026-01-01T00:00:00Z") == 2
